"""One-time/periodic offline ETL: DOL OFLC LCA disclosure data -> `SponsorHistory`.

Not a Lambda - run manually (or from a periodic task outside the hourly scan
loop) whenever a new DOL OFLC quarterly disclosure file is published. Get
the current file from the DOL OFLC "Performance Data" page:

    https://www.dol.gov/agencies/eta/foreign-labor/performance

Disclosure files are published there as dated CSV/XLSX downloads per fiscal
year/quarter; the exact filename and URL change every release, so this
script takes a local file path you've already downloaded rather than a
hardcoded (and inevitably stale) download URL.

Usage:
    python -m pipeline.dol_lca_etl --input LCA_Disclosure_Data_FY2026_Q2.xlsx

Only rows with CASE_STATUS "Certified" are loaded - a denied or withdrawn
filing isn't evidence the employer actually sponsors, and rows older than
`--lookback-years` (default: `LCA_LOOKBACK_YEARS`) are dropped at ETL time
so `SponsorHistory` never needs its own separate pruning job.

Known open problem (see docs/ARCHITECTURE.md "Visa + experience-level
filtering"): DOL filer legal-entity names often differ from the brand name
on a job posting. This script keys `employer_normalized` off the DBA/trade
name column when the disclosure data provides one (closer to a brand name
than the legal filer name), falling back to the legal employer name
otherwise; `employer_normalize.py`'s curated alias table plus
`visa_level_filter.fetch_sponsor_filings`'s fuzzy-match fallback handle the
rest at query time.
"""

from __future__ import annotations

import argparse
import csv
from collections.abc import Iterable, Iterator
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from pipeline.config import LCA_LOOKBACK_YEARS
from pipeline.dynamo import get_table
from pipeline.employer_normalize import known_employer_aliases, normalize_employer_name
from shared.tables import SPONSOR_HISTORY_TABLE

# DOL has renamed several of these columns across fiscal-year vintages of the
# disclosure file; each tuple lists known variants, checked in order, so this
# script tolerates a schema change without needing an edit for every field.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "case_id": ("CASE_NUMBER",),
    "case_status": ("CASE_STATUS",),
    "decision_date": ("DECISION_DATE",),
    "employer_name": ("EMPLOYER_NAME",),
    "employer_dba": ("EMPLOYER_BUSINESS_DBA", "TRADE_NAME_DBA"),
    "job_title": ("JOB_TITLE",),
    "soc_title": ("SOC_TITLE",),
    "pw_wage_level": ("PW_WAGE_LEVEL", "PREVAILING_WAGE_LEVEL", "WAGE_LEVEL"),
    "wage_from": ("WAGE_RATE_OF_PAY_FROM_1", "WAGE_RATE_OF_PAY_FROM", "WAGE_FROM"),
    "wage_to": ("WAGE_RATE_OF_PAY_TO_1", "WAGE_RATE_OF_PAY_TO", "WAGE_TO"),
    "wage_unit": ("WAGE_UNIT_OF_PAY_1", "WAGE_UNIT_OF_PAY", "WAGE_UNIT"),
}

_WAGE_UNIT_ANNUALIZE: dict[str, float] = {
    "year": 1,
    "yr": 1,
    "month": 12,
    "mth": 12,
    "bi-weekly": 26,
    "biweekly": 26,
    "week": 52,
    "wk": 52,
    "hour": 2080,
    "hr": 2080,
}

_DATE_FORMATS = ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y")


def _get_col(row: dict[str, str], field: str) -> str | None:
    for alias in COLUMN_ALIASES[field]:
        value = row.get(alias)
        if value not in (None, ""):
            return str(value)
    return None


def _parse_date(value: str) -> date | None:
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _annualize_wage(value: str | None, unit: str | None) -> float | None:
    if not value:
        return None
    try:
        amount = float(str(value).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None
    multiplier = _WAGE_UNIT_ANNUALIZE.get((unit or "year").strip().lower(), 1)
    return amount * multiplier


def row_to_sponsor_record(row: dict[str, str], *, cutoff: date) -> dict | None:
    """Convert one raw disclosure-file row into a `SponsorHistory` item, or
    None if the row should be dropped (not certified, outside the lookback
    window, or missing required fields).
    """
    case_status = (_get_col(row, "case_status") or "").strip().lower()
    if case_status and case_status != "certified":
        return None

    decision_date = _parse_date(_get_col(row, "decision_date") or "")
    if decision_date is None or decision_date < cutoff:
        return None

    case_id = _get_col(row, "case_id")
    employer_name = _get_col(row, "employer_name")
    if not case_id or not employer_name:
        return None

    employer_key_source = _get_col(row, "employer_dba") or employer_name
    employer_normalized = normalize_employer_name(employer_key_source)

    wage_unit = _get_col(row, "wage_unit")
    wage_from = _annualize_wage(_get_col(row, "wage_from"), wage_unit)
    wage_to = _annualize_wage(_get_col(row, "wage_to"), wage_unit)

    return {
        "employer_normalized": employer_normalized,
        "decision_date_case_id": f"{decision_date.isoformat()}#{case_id}",
        "job_title": _get_col(row, "job_title") or "",
        "soc_title": _get_col(row, "soc_title") or "",
        "wage_from": Decimal(str(wage_from)) if wage_from is not None else None,
        "wage_to": Decimal(str(wage_to)) if wage_to is not None else None,
        "pw_wage_level": _get_col(row, "pw_wage_level"),
    }


# --- File readers ---


def iter_rows_csv(path: Path) -> Iterator[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        yield from csv.DictReader(f)


def iter_rows_xlsx(path: Path) -> Iterator[dict[str, str]]:
    import openpyxl  # deferred: only needed for the (large) XLSX path

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        for row in rows:
            yield {header[i]: row[i] for i in range(len(header)) if i < len(row)}
    finally:
        workbook.close()


def iter_rows(path: Path) -> Iterator[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return iter_rows_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return iter_rows_xlsx(path)
    raise ValueError(f"Unsupported DOL disclosure file type: {suffix!r} (expected .csv or .xlsx)")


# --- Load ---


def load_sponsor_history(records: Iterable[dict]) -> int:
    table = get_table(SPONSOR_HISTORY_TABLE)
    count = 0
    with table.batch_writer(overwrite_by_pkeys=["employer_normalized", "decision_date_case_id"]) as batch:
        for record in records:
            batch.put_item(Item=record)
            count += 1
    return count


def resolve_employer_filter(brands: list[str]) -> set[str]:
    """Expand `--employers` brand names (e.g. "amazon") into the full set of
    normalized filer-name variants from `employer_normalize.py`'s curated
    alias table, per docs/ARCHITECTURE.md - this only needs to cover
    companies actively targeted by the job search, not the whole dataset.
    National LCA disclosure files run into the hundreds of thousands of
    rows; `SponsorHistory` is provisioned at 1 WCU (to fit AWS's Always-Free
    tier - see infra/stacks/data_stack.py), so loading the unfiltered file
    would take an unreasonable amount of time. Scoping to known target
    employers up front is the fix, not a workaround.
    """
    allowed: set[str] = set()
    for brand in brands:
        normalized_brand = normalize_employer_name(brand)
        allowed.add(normalized_brand)
        allowed.update(known_employer_aliases(normalized_brand))
    return allowed


def run_etl(
    input_path: Path,
    *,
    lookback_years: int = LCA_LOOKBACK_YEARS,
    as_of: date | None = None,
    employer_filter: set[str] | None = None,
) -> int:
    as_of = as_of or date.today()
    cutoff = as_of.replace(year=as_of.year - lookback_years)
    records = (
        record
        for row in iter_rows(input_path)
        if (record := row_to_sponsor_record(row, cutoff=cutoff)) is not None
        and (employer_filter is None or record["employer_normalized"] in employer_filter)
    )
    return load_sponsor_history(records)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", required=True, type=Path, help="Local DOL LCA disclosure CSV/XLSX file")
    parser.add_argument("--lookback-years", type=int, default=LCA_LOOKBACK_YEARS)
    parser.add_argument(
        "--employers",
        nargs="+",
        default=None,
        help=(
            "Restrict load to these brands' known filer-name aliases (e.g. "
            "--employers amazon google). Omit to load every employer in the "
            "file - not recommended given SponsorHistory's 1-WCU provisioned "
            "capacity; see resolve_employer_filter's docstring."
        ),
    )
    args = parser.parse_args(argv)

    employer_filter = resolve_employer_filter(args.employers) if args.employers else None
    count = run_etl(args.input, lookback_years=args.lookback_years, employer_filter=employer_filter)
    print(f"Loaded {count} SponsorHistory records from {args.input}")


if __name__ == "__main__":
    main()
